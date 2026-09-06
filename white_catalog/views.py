import hashlib
import json
import logging
from functools import wraps
from decimal import Decimal
from io import BytesIO
from urllib.parse import urlencode
from django.conf import settings
from django.contrib.auth import authenticate
from django.shortcuts import render, get_object_or_404, redirect
from django.contrib import messages
from django.http import JsonResponse, HttpResponse
from django.urls import reverse
from django.utils import timezone
from django.utils.html import strip_tags
from django.views.decorators.http import require_POST
from .models import (
    WhiteCategory, WhiteSubcategory, WhiteCatalogUser, WhiteCatalogUserActivity,
    WhiteFabricType, WhiteProductVariant, WhiteVariantPackPrice, WhitePackType,
    WhiteColorVariant,
    WhiteCart, WhiteCartItem, WhiteOrder, WhiteOrderItem,
    apply_price_list,
)
from .middleware import get_client_ip


logger = logging.getLogger(__name__)


# ---------------------------------------------------------------------------
# Auth helpers
# ---------------------------------------------------------------------------

def _catalog_session_login(request, user):
    """Persist the white catalog session for the selected user."""
    request.session["white_catalog_user_id"] = user.id
    request.session["white_catalog_username"] = user.username
    request.session["white_catalog_company_name"] = user.company_name


def _sync_admin_to_catalog_user(django_user, raw_password=None):
    """Map a Django staff/superuser account into a WhiteCatalogUser account."""
    if not getattr(django_user, "is_authenticated", False):
        return None
    if not django_user.is_active or not (django_user.is_staff or django_user.is_superuser):
        return None

    full_name = (
        django_user.get_full_name().strip()
        or getattr(django_user, "first_name", "").strip()
        or django_user.get_username()
    )
    defaults = {
        "company_name": "Admin Test Account",
        "contact_name": full_name,
        "contact_phone": "",
        "is_active": True,
    }
    catalog_user, created = WhiteCatalogUser.objects.get_or_create(
        username=django_user.get_username(),
        defaults=defaults,
    )

    update_fields = []
    if not catalog_user.company_name:
        catalog_user.company_name = defaults["company_name"]
        update_fields.append("company_name")
    if not catalog_user.contact_name:
        catalog_user.contact_name = full_name
        update_fields.append("contact_name")
    if catalog_user.contact_phone is None:
        catalog_user.contact_phone = ""
        update_fields.append("contact_phone")
    if not catalog_user.is_active:
        catalog_user.is_active = True
        update_fields.append("is_active")
    if raw_password:
        catalog_user.set_password(raw_password)
        update_fields.append("password_hash")

    if created or update_fields:
        catalog_user.save(update_fields=update_fields or None)

    return catalog_user


def get_current_catalog_user(request):
    """Return the logged-in WhiteCatalogUser or None."""
    user_id = request.session.get("white_catalog_user_id")
    if not user_id:
        admin_catalog_user = _sync_admin_to_catalog_user(getattr(request, "user", None))
        if admin_catalog_user:
            _catalog_session_login(request, admin_catalog_user)
            return admin_catalog_user
        return None
    try:
        return WhiteCatalogUser.objects.get(pk=user_id, is_active=True)
    except WhiteCatalogUser.DoesNotExist:
        admin_catalog_user = _sync_admin_to_catalog_user(getattr(request, "user", None))
        if admin_catalog_user:
            _catalog_session_login(request, admin_catalog_user)
            return admin_catalog_user
        return None


def require_catalog_login(view_func):
    """Decorator: redirect unauthenticated users to login page."""
    @wraps(view_func)
    def wrapper(request, *args, **kwargs):
        if not get_current_catalog_user(request):
            return redirect(f"/white-catalog/login/?next={request.path}")
        return view_func(request, *args, **kwargs)
    return wrapper


def get_or_create_active_cart(user):
    """Return the user's active cart, creating one if needed."""
    cart, _ = WhiteCart.objects.get_or_create(user=user, status=WhiteCart.STATUS_ACTIVE)
    return cart


def _build_cart_page_context(request):
    """Cart, grouped line items, and count for cart page / drawer."""
    user = get_current_catalog_user(request)
    if not user:
        return {"cart": None, "grouped_items": [], "cart_count": 0}
    try:
        cart = WhiteCart.objects.prefetch_related(
            "items__product__images", "items__variant__fabric_type", "items__variant__size_type",
            "items__color_variant__color", "items__pack_type"
        ).get(user=user, status=WhiteCart.STATUS_ACTIVE)
    except WhiteCart.DoesNotExist:
        cart = None

    grouped_items = []
    if cart:
        groups_map = {}
        for item in cart.items.all():
            if item.color_variant_id:
                key = ("color", item.product_id)
            elif not item.variant_id or not item.pack_type_id:
                key = ("simple", item.product_id)
            else:
                key = (item.product_id, item.variant.fabric_type_id, item.pack_type_id)
            if key not in groups_map:
                group = {
                    "product": item.product,
                    "product_id": item.product_id,
                    "product_name": item.product.name,
                    "product_image": item.product.get_main_image(),
                    "fabric_type": item.display_fabric_name,
                    "pack_type": item.display_pack_name,
                    "price_at_add": item.price_at_add,
                    "items": [],
                    "group_total": Decimal("0"),
                    "items_by_variant_id": {},
                    "is_simple": item.is_simple_item,
                    "is_color": bool(item.color_variant_id),
                }
                if item.variant_id and item.pack_type_id:
                    group["pack_type_obj"] = item.pack_type
                    group["fabric_type_id"] = item.variant.fabric_type_id
                groups_map[key] = group
                grouped_items.append(group)
            groups_map[key]["items_by_variant_id"][
                item.color_variant_id or item.variant_id or f"simple-{item.product_id}"
            ] = item
            groups_map[key]["group_total"] += item.price_at_add * item.quantity

        for group in grouped_items:
            if group["is_color"]:
                color_variants = (
                    group["product"].color_variants.filter(is_active=True)
                    .select_related("color")
                )
                for cv in color_variants:
                    existing_item = group["items_by_variant_id"].get(cv.id)
                    quantity = existing_item.quantity if existing_item else 0
                    line_total = existing_item.get_line_total() if existing_item else Decimal("0")
                    group["items"].append(
                        {
                            "item_id": existing_item.id if existing_item else "",
                            "variant_id": "",
                            "color_variant_id": cv.id,
                            "pack_type_id": "",
                            "barcode": cv.barcode or "",
                            "size_name": cv.color.name,
                            "color_hex": cv.color.hex_color or "",
                            "swatch_url": cv.color.swatch_image.url if cv.color.swatch_image else "",
                            "quantity": quantity,
                            "line_total": "{:.2f}".format(line_total),
                            "has_item": bool(existing_item),
                            "is_simple": False,
                            "is_color": True,
                        }
                    )
            elif group["is_simple"]:
                existing_item = group["items_by_variant_id"].get(f"simple-{group['product'].id}")
                group["items"].append(
                    {
                        "item_id": existing_item.id if existing_item else "",
                        "variant_id": "",
                        "pack_type_id": "",
                        "barcode": "",
                        "size_name": "יחידות",
                        "quantity": existing_item.quantity if existing_item else 0,
                        "line_total": "{:.2f}".format(existing_item.get_line_total() if existing_item else Decimal('0')),
                        "has_item": bool(existing_item),
                        "is_simple": True,
                    }
                )
            else:
                variants = (
                    group["product"].variants.filter(
                        is_active=True,
                        fabric_type_id=group["fabric_type_id"],
                    )
                    .select_related("size_type")
                    .order_by("size_type_id", "id")
                )
                for variant in variants:
                    existing_item = group["items_by_variant_id"].get(variant.id)
                    quantity = existing_item.quantity if existing_item else 0
                    line_total = existing_item.get_line_total() if existing_item else Decimal("0")
                    group["items"].append(
                        {
                            "item_id": existing_item.id if existing_item else "",
                            "variant_id": variant.id,
                            "pack_type_id": group["pack_type_obj"].id,
                            "barcode": variant.barcode or "",
                            "size_name": variant.size_type.name,
                            "quantity": quantity,
                            "line_total": "{:.2f}".format(line_total),
                            "has_item": bool(existing_item),
                            "is_simple": False,
                        }
                    )
            del group["items_by_variant_id"]
            del group["product"]
            if "pack_type_obj" in group:
                del group["pack_type_obj"]
            if "fabric_type_id" in group:
                del group["fabric_type_id"]

    return {
        "cart": cart,
        "grouped_items": grouped_items,
        "cart_count": cart.get_item_count() if cart else 0,
    }


def _build_order_grouped_items(order):
    """Group submitted order items like the cart drawer summary."""
    grouped_items = []
    groups_map = {}

    for item in order.items.all():
        is_color = bool(item.color_name) or bool(item.color_variant_id)
        is_simple = not item.variant_id and not is_color
        if is_color:
            key = ("color", item.product_id or item.product_name)
        elif is_simple:
            key = ("simple", item.product_id or item.product_name)
        else:
            fabric_name = item.variant.fabric_type.name if item.variant_id and item.variant and item.variant.fabric_type_id else item.variant_name
            key = (item.product_id or item.product_name, fabric_name, item.pack_type_name)

        if key not in groups_map:
            product_image = item.product.get_main_image() if item.product_id and item.product else None
            if is_color:
                fabric_label = "מניפת צבעים"
            elif is_simple:
                fabric_label = "ללא גרסאות"
            else:
                fabric_label = item.variant.fabric_type.name if item.variant_id and item.variant and item.variant.fabric_type_id else item.variant_name
            group = {
                "product_name": item.product_name,
                "product_image": product_image,
                "fabric_type": fabric_label,
                "pack_type": item.pack_type_name,
                "price_at_add": item.unit_price,
                "items": [],
                "group_total": Decimal("0"),
                "is_simple": is_simple,
                "is_color": is_color,
            }
            groups_map[key] = group
            grouped_items.append(group)

        if is_color:
            row_name = item.color_name or item.variant_name
        elif is_simple:
            row_name = "יחידות"
        else:
            row_name = item.size_name
        groups_map[key]["items"].append(
            {
                "size_name": row_name,
                "barcode": item.barcode or "",
                "quantity": item.quantity,
                "line_total": "{:.2f}".format(item.get_line_total()),
            }
        )
        groups_map[key]["group_total"] += item.get_line_total()

    return grouped_items


def _nav_context(request=None):
    """Common navigation context shared by all views."""
    context = {
        "all_categories": WhiteCategory.objects.all(),
        "standalone_subcategories": WhiteSubcategory.objects.filter(category__isnull=True),
    }
    if request is not None:
        context["catalog_user"] = get_current_catalog_user(request)
    return context


def _cart_count(request):
    user = get_current_catalog_user(request)
    if not user:
        return 0
    try:
        cart = WhiteCart.objects.get(user=user, status=WhiteCart.STATUS_ACTIVE)
        return cart.get_item_count()
    except WhiteCart.DoesNotExist:
        return 0


def catalog_home(request):
    """Main white catalog page showing all categories and standalone subcategories."""
    categories = WhiteCategory.objects.filter(show_products_on_homepage=False)
    standalone_subcategories = WhiteSubcategory.objects.filter(category__isnull=True).prefetch_related("images")
    category_homepage_subcategories = WhiteSubcategory.objects.filter(
        category__show_products_on_homepage=True
    ).select_related("category").prefetch_related("images").order_by(
        "category__order", "category__name", "order", "name"
    )
    context = {
        "categories": categories,
        "standalone_subcategories": standalone_subcategories,
        "homepage_subcategories": list(category_homepage_subcategories) + list(standalone_subcategories),
        "all_categories": WhiteCategory.objects.all(),
        "catalog_user": get_current_catalog_user(request),
        "cart_count": _cart_count(request),
    }
    return render(request, "white_catalog/catalog_home.html", context)


def barcode_search(request):
    """Find a white-catalog variant by barcode and redirect to its product page."""
    q = (request.GET.get("q") or request.POST.get("q") or "").strip()
    if not q:
        messages.error(request, "נא להזין ברקוד לחיפוש")
        return redirect("white_catalog:home")

    variant = (
        WhiteProductVariant.objects.filter(barcode__iexact=q, is_active=True)
        .select_related("product", "product__category", "fabric_type", "size_type")
        .first()
    )

    color_variant = None
    if not variant or not variant.product_id:
        color_variant = (
            WhiteColorVariant.objects.filter(barcode__iexact=q, is_active=True)
            .select_related("product", "product__category", "color")
            .first()
        )

    if variant and variant.product_id:
        product = variant.product
        query = urlencode({
            "fabric": variant.fabric_type_id,
            "variant": variant.id,
        })
    elif color_variant and color_variant.product_id:
        product = color_variant.product
        query = urlencode({"color_variant": color_variant.id})
    else:
        messages.error(request, f'לא נמצא מוצר עם ברקוד "{q}"')
        referer = request.META.get("HTTP_REFERER")
        if referer:
            return redirect(referer)
        return redirect("white_catalog:home")

    if product.category_id:
        url = reverse(
            "white_catalog:subcategory_detail",
            kwargs={
                "category_slug": product.category.slug,
                "subcategory_slug": product.slug,
            },
        )
    else:
        url = reverse(
            "white_catalog:standalone_subcategory_detail",
            kwargs={"subcategory_slug": product.slug},
        )
    return redirect(f"{url}?{query}")


def category_detail(request, category_slug):
    """Category detail page showing subcategories."""
    category = get_object_or_404(WhiteCategory, slug=category_slug)
    context = {
        **_nav_context(request),
        "category": category,
        "subcategories": category.subcategories.all(),
        "catalog_user": get_current_catalog_user(request),
        "cart_count": _cart_count(request),
    }
    return render(request, "white_catalog/category_detail.html", context)


def _subcategory_detail_context(request, subcategory, category=None):
    """Build context for subcategory detail views."""
    catalog_user = get_current_catalog_user(request)
    show_variant_ordering = bool(catalog_user and subcategory.is_orderable and subcategory.has_order_variants)
    show_color_ordering = bool(
        catalog_user and subcategory.is_orderable
        and subcategory.has_color_variants and not subcategory.has_order_variants
    )
    show_simple_ordering = bool(
        catalog_user and subcategory.is_orderable
        and not subcategory.has_order_variants and not subcategory.has_color_variants
    )

    # Build variants JSON grouped by fabric_type for the ordering widget
    variants_data = []
    pack_types_data = []
    color_variants_data = []
    simple_cart_quantity = 0
    active_cart = None
    if catalog_user:
        try:
            active_cart = WhiteCart.objects.prefetch_related("items").get(
                user=catalog_user,
                status=WhiteCart.STATUS_ACTIVE,
            )
        except WhiteCart.DoesNotExist:
            active_cart = None

    if show_variant_ordering:
        cart_qty_map = {}
        if active_cart:
            cart_qty_map = {
                (item.variant_id, item.pack_type_id): item.quantity
                for item in active_cart.items.filter(
                    product=subcategory,
                    variant__isnull=False,
                    pack_type__isnull=False,
                )
            }

        # Restrict to the pack types this user's route allows.
        allowed_pack_ids = set(catalog_user.get_allowed_pack_types().values_list("id", flat=True))

        pack_types_data = [
            {"pack_id": pt.pk, "pack_name": pt.name, "pack_qty": pt.quantity}
            for pt in WhitePackType.objects.filter(is_active=True)
            if pt.pk in allowed_pack_ids
        ]

        fabric_map = {}
        for variant in (subcategory.variants
                        .filter(is_active=True)
                        .select_related("fabric_type", "size_type")
                        .prefetch_related("pack_types")):
            # A variant with no pack types is treated as available in all allowed packs.
            variant_pack_ids = {pt.pk for pt in variant.pack_types.all()}
            if variant_pack_ids:
                size_pack_ids = variant_pack_ids & allowed_pack_ids
            else:
                size_pack_ids = set(allowed_pack_ids)
            if not size_pack_ids:
                continue

            fid = variant.fabric_type_id
            if fid not in fabric_map:
                fabric_map[fid] = {
                    "fabric_id": fid,
                    "name": variant.fabric_type.name,
                    "sizes": [],
                }
            raw_unit = variant.unit_price if variant.unit_price is not None else subcategory.unit_price
            list_price = apply_price_list(raw_unit, None)
            effective_price = apply_price_list(raw_unit, catalog_user)
            fabric_map[fid]["sizes"].append({
                "variant_id": variant.id,
                "size_name": variant.size_type.name,
                "unit_price": str(effective_price) if effective_price is not None else None,
                "list_unit_price": str(list_price) if list_price is not None else None,
                "barcode": variant.barcode or "",
                "pack_ids": sorted(size_pack_ids),
                "cart_quantities": {
                    str(pt["pack_id"]): cart_qty_map.get((variant.id, pt["pack_id"]), 0)
                    for pt in pack_types_data
                },
            })
        variants_data = list(fabric_map.values())
    elif show_color_ordering:
        cart_qty_map = {}
        if active_cart:
            cart_qty_map = {
                item.color_variant_id: item.quantity
                for item in active_cart.items.filter(
                    product=subcategory,
                    color_variant__isnull=False,
                )
            }

        for cv in (subcategory.color_variants
                   .filter(is_active=True, color__is_active=True)
                   .select_related("color")):
            raw_unit = cv.get_effective_price()
            list_price = apply_price_list(raw_unit, None)
            effective_price = apply_price_list(raw_unit, catalog_user)
            color_variants_data.append({
                "id": cv.id,
                "color_name": cv.color.name,
                "hex_color": cv.color.hex_color or "",
                "swatch_url": cv.color.swatch_image.url if cv.color.swatch_image else "",
                "image_url": cv.image.url if cv.image else "",
                "barcode": cv.barcode or "",
                "unit_price": str(effective_price) if effective_price is not None else None,
                "list_unit_price": str(list_price) if list_price is not None else None,
                "cart_quantity": cart_qty_map.get(cv.id, 0),
            })
    elif show_simple_ordering and active_cart:
        simple_item = (
            active_cart.items.filter(
                product=subcategory,
                variant__isnull=True,
                pack_type__isnull=True,
            )
            .first()
        )
        if simple_item:
            simple_cart_quantity = simple_item.quantity

    return {
        **_nav_context(request),
        "category": category,
        "subcategory": subcategory,
        "subcategory_images": subcategory.images.all(),
        "gallery_images": _product_gallery_images(subcategory),
        "catalog_user": catalog_user,
        "show_variant_ordering": show_variant_ordering,
        "show_color_ordering": show_color_ordering,
        "show_simple_ordering": show_simple_ordering,
        "variants_data": variants_data,
        "pack_types_data": pack_types_data,
        "color_variants_data": color_variants_data,
        "color_variants_have_images": any(
            (cv.get("image_url") or cv.get("swatch_url")) for cv in color_variants_data
        ),
        "simple_cart_quantity": simple_cart_quantity,
        "display_unit_price": apply_price_list(subcategory.unit_price, catalog_user),
        "display_list_unit_price": apply_price_list(subcategory.unit_price, None),
        "cart_count": _cart_count(request),
    }


LEGACY_PRODUCT_SLUGS = {
    "מארז-חיתולי-טטרה-מודפס-3": ("baby_diapers", "muslin_baby_diapers"),
}


def _image_fingerprint(file_field):
    """Identify the same photo even when stored as different files/encodings."""
    if not file_field:
        return None
    try:
        file_field.open("rb")
        data = file_field.read()
        file_field.close()
    except Exception:
        return getattr(file_field, "url", None) or getattr(file_field, "name", None)
    if not data:
        return getattr(file_field, "url", None) or getattr(file_field, "name", None)
    try:
        from PIL import Image
        img = Image.open(BytesIO(data)).convert("L").resize((12, 12))
        pixels = list(img.getdata())
        avg = sum(pixels) / len(pixels) if pixels else 0
        return "".join("1" if pixel >= avg else "0" for pixel in pixels)
    except Exception:
        return hashlib.md5(data).hexdigest()


def _product_gallery_images(subcategory):
    """Main image + gallery + print/color photos, unique by image content."""
    images = []
    seen = set()

    def add(file_field, alt):
        if not file_field:
            return
        url = getattr(file_field, "url", "") or ""
        if not url:
            return
        key = _image_fingerprint(file_field) or url
        if key in seen:
            return
        seen.add(key)
        images.append({"url": url, "alt": alt})

    if subcategory.image:
        add(subcategory.image, subcategory.name)
    for img in subcategory.images.all():
        add(img.image, img.alt_text or subcategory.name)
    for cv in (subcategory.color_variants
               .filter(is_active=True)
               .select_related("color")):
        add(cv.image, cv.color.name)
    return images


def subcategory_detail(request, category_slug, subcategory_slug):
    """Subcategory detail page with image gallery."""
    legacy = LEGACY_PRODUCT_SLUGS.get(subcategory_slug)
    if legacy:
        return redirect(
            "white_catalog:subcategory_detail",
            category_slug=legacy[0],
            subcategory_slug=legacy[1],
            permanent=True,
        )
    category = get_object_or_404(WhiteCategory, slug=category_slug)
    subcategory = get_object_or_404(WhiteSubcategory, category=category, slug=subcategory_slug)
    return render(request, "white_catalog/subcategory_detail.html",
                  _subcategory_detail_context(request, subcategory, category))


def standalone_subcategory_detail(request, subcategory_slug):
    """Standalone subcategory detail page (without category)."""
    subcategory = get_object_or_404(WhiteSubcategory, slug=subcategory_slug, category__isnull=True)
    return render(request, "white_catalog/subcategory_detail.html",
                  _subcategory_detail_context(request, subcategory))


def login_view(request):
	"""Login view for white catalog users."""
	if request.method == "POST":
		username = request.POST.get("username", "").strip()
		password = request.POST.get("password", "")
		
		if not username or not password:
			messages.error(request, "נא למלא שם משתמש וסיסמא")
		else:
			try:
				user = WhiteCatalogUser.objects.get(username=username, is_active=True)
				password_ok = user.check_password(password)
			except WhiteCatalogUser.DoesNotExist:
				user = None
				password_ok = False

			if not password_ok:
				django_user = authenticate(request, username=username, password=password)
				if django_user and django_user.is_active and (django_user.is_staff or django_user.is_superuser):
					user = _sync_admin_to_catalog_user(django_user, raw_password=password)
					password_ok = bool(user)

			if password_ok and user:
				_catalog_session_login(request, user)

				# Update last login time
				user.last_login = timezone.now()
				user.save(update_fields=['last_login'])

				# Log activity
				try:
					WhiteCatalogUserActivity.objects.create(
						user=user,
						ip_address=get_client_ip(request),
						user_agent=request.META.get('HTTP_USER_AGENT', '')[:500],
						page_url=request.path[:500]
					)
				except Exception:
					pass  # Don't let logging errors break the login flow

				# Update activity timestamp
				user.update_activity()

				messages.success(request, f"ברוך הבא, {user.company_name}!")

				# Redirect to next page or home
				next_url = request.GET.get("next") or request.POST.get("next") or "white_catalog:home"
				return redirect(next_url)

			messages.error(request, "שם משתמש או סיסמא שגויים")
	
	# Get all categories for navigation
	all_categories = WhiteCategory.objects.all()
	standalone_subcategories = WhiteSubcategory.objects.filter(category__isnull=True)
	
	context = {
		"all_categories": all_categories,
		"standalone_subcategories": standalone_subcategories,
	}
	return render(request, "white_catalog/login.html", context)


def logout_view(request):
    """Logout view for white catalog users."""
    for key in ("white_catalog_user_id", "white_catalog_username", "white_catalog_company_name"):
        request.session.pop(key, None)
    messages.success(request, "התנתקת בהצלחה")
    return redirect("white_catalog:home")


# ---------------------------------------------------------------------------
# Cart views
# ---------------------------------------------------------------------------

@require_POST
@require_catalog_login
def cart_add(request):
    """Add/update items in the cart from the product ordering form."""
    user = get_current_catalog_user(request)
    cart = get_or_create_active_cart(user)

    want_json = request.POST.get("format") == "json"
    product_id = request.POST.get("product_id")
    simple_quantity_raw = request.POST.get("simple_quantity")

    if product_id and simple_quantity_raw is not None:
        try:
            product = WhiteSubcategory.objects.get(pk=product_id, is_orderable=True)
        except (WhiteSubcategory.DoesNotExist, TypeError, ValueError):
            if want_json:
                return JsonResponse({"ok": False, "error": "המוצר לא זמין להזמנה"}, status=400)
            messages.error(request, "המוצר לא זמין להזמנה")
            return redirect(request.POST.get("next") or "white_catalog:cart")

        if product.has_order_variants:
            if want_json:
                return JsonResponse({"ok": False, "error": "למוצר זה יש להזין הזמנה לפי גרסאות ומארזים"}, status=400)
            messages.error(request, "למוצר זה יש להזין הזמנה לפי גרסאות ומארזים")
            return redirect(request.POST.get("next") or "white_catalog:cart")

        if product.has_color_variants:
            if want_json:
                return JsonResponse({"ok": False, "error": "למוצר זה יש להזין הזמנה לפי מניפת הצבעים"}, status=400)
            messages.error(request, "למוצר זה יש להזין הזמנה לפי מניפת הצבעים")
            return redirect(request.POST.get("next") or "white_catalog:cart")

        try:
            quantity = int((simple_quantity_raw or "").strip() or 0)
        except (TypeError, ValueError, AttributeError):
            quantity = 0

        if quantity < 0:
            quantity = 0

        listed_price = apply_price_list(product.unit_price, user)
        if listed_price is None:
            if want_json:
                return JsonResponse({"ok": False, "error": "מחיר לא זמין"}, status=400)
            messages.error(request, "מחיר לא זמין")
            return redirect(request.POST.get("next") or "white_catalog:cart")

        existing_item = (
            WhiteCartItem.objects.filter(
                cart=cart,
                product=product,
                variant__isnull=True,
                pack_type__isnull=True,
                color_variant__isnull=True,
            )
            .first()
        )

        added = 1 if quantity > 0 else 0
        if quantity == 0:
            if existing_item:
                existing_item.delete()
        elif existing_item:
            existing_item.quantity = quantity
            existing_item.price_at_add = listed_price
            existing_item.save(update_fields=["quantity", "price_at_add", "updated"])
        else:
            WhiteCartItem.objects.create(
                cart=cart,
                product=product,
                quantity=quantity,
                price_at_add=listed_price,
            )

        if want_json:
            return JsonResponse(
                {
                    "ok": True,
                    "added": added,
                    "cart_count": cart.get_item_count(),
                    "message": (
                        "דף ההזמנה עודכן"
                        if added
                        else "המוצר הוסר מדף ההזמנה"
                    ),
                }
            )

        if added:
            messages.success(request, "דף ההזמנה עודכן")
        else:
            messages.info(request, "המוצר הוסר מדף ההזמנה")
        return redirect(request.POST.get("next") or "white_catalog:cart")

    # Color-fan ordering: one quantity per color variant, no pack types.
    if any(key.startswith("color_qty_") for key in request.POST):
        added = 0
        for key, value in request.POST.items():
            if not key.startswith("color_qty_"):
                continue
            try:
                color_variant_id = key[len("color_qty_"):]
                raw = (value or "").strip()
                quantity = 0 if raw == "" else int(raw)
            except (ValueError, TypeError, AttributeError):
                continue

            if quantity < 0:
                continue

            try:
                color_variant = WhiteColorVariant.objects.select_related("product").get(
                    pk=color_variant_id, is_active=True
                )
            except (WhiteColorVariant.DoesNotExist, ValueError):
                continue

            if not color_variant.product.is_orderable or not color_variant.product.has_color_variants:
                continue

            effective_price = apply_price_list(color_variant.get_effective_price(), user)
            if effective_price is None:
                continue

            if quantity == 0:
                WhiteCartItem.objects.filter(cart=cart, color_variant=color_variant).delete()
            else:
                WhiteCartItem.objects.update_or_create(
                    cart=cart,
                    color_variant=color_variant,
                    defaults={
                        "product": color_variant.product,
                        "quantity": quantity,
                        "price_at_add": effective_price,
                    },
                )
                added += 1

        if want_json:
            return JsonResponse(
                {
                    "ok": True,
                    "added": added,
                    "cart_count": cart.get_item_count(),
                    "message": (
                        f"דף ההזמנה עודכן — {added} צבעים נוספו"
                        if added
                        else "לא נוספו פריטים להזמנה"
                    ),
                }
            )

        if added:
            messages.success(request, f"דף ההזמנה עודכן — {added} צבעים נוספו")
        else:
            messages.info(request, "לא נוספו פריטים להזמנה")

        return redirect(request.POST.get("next") or "white_catalog:cart")

    pack_type_id = request.POST.get("pack_type_id")
    try:
        pack_type = WhitePackType.objects.get(pk=pack_type_id, is_active=True)
    except (WhitePackType.DoesNotExist, TypeError, ValueError):
        if want_json:
            return JsonResponse({"ok": False, "error": "נא לבחור סוג אריזה"}, status=400)
        messages.error(request, "נא לבחור סוג אריזה")
        return redirect(request.POST.get("next") or "white_catalog:cart")

    if pack_type.quantity != 3:
        if want_json:
            return JsonResponse({"ok": False, "error": "ביגוד מוזמן בשלישיות בלבד"}, status=400)
        messages.error(request, "ביגוד מוזמן בשלישיות בלבד")
        return redirect(request.POST.get("next") or "white_catalog:cart")

    # Enforce the user's pack route — they may only order allowed pack forms.
    allowed_pack_ids = set(user.get_allowed_pack_types().values_list("id", flat=True))
    if pack_type.pk not in allowed_pack_ids:
        if want_json:
            return JsonResponse({"ok": False, "error": "סוג האריזה אינו זמין עבורך"}, status=400)
        messages.error(request, "סוג האריזה אינו זמין עבורך")
        return redirect(request.POST.get("next") or "white_catalog:cart")

    added = 0
    for key, value in request.POST.items():
        if not key.startswith("qty_"):
            continue
        try:
            variant_id = key[4:]
            raw = (value or "").strip()
            quantity = 0 if raw == "" else int(raw)
        except (ValueError, TypeError, AttributeError):
            continue

        if quantity < 0:
            continue

        try:
            variant = WhiteProductVariant.objects.select_related("product").get(
                pk=variant_id, is_active=True
            )
        except WhiteProductVariant.DoesNotExist:
            continue

        if not variant.product.is_orderable or not variant.product.has_order_variants:
            continue

        # Skip variants that aren't sold in the selected pack form (empty = all forms).
        variant_pack_ids = set(variant.pack_types.values_list("id", flat=True))
        if variant_pack_ids and pack_type.pk not in variant_pack_ids:
            continue

        effective_price = apply_price_list(
            variant.unit_price if variant.unit_price is not None else variant.product.unit_price,
            user,
        )
        if effective_price is None:
            continue

        # Price per pack = listed unit_price × units in pack
        price_per_pack = effective_price * pack_type.quantity

        if quantity == 0:
            WhiteCartItem.objects.filter(cart=cart, variant=variant, pack_type=pack_type).delete()
        else:
            WhiteCartItem.objects.update_or_create(
                cart=cart,
                variant=variant,
                pack_type=pack_type,
                defaults={
                    "product": variant.product,
                    "quantity": quantity,
                    "price_at_add": price_per_pack,
                }
            )
            added += 1

    if want_json:
        return JsonResponse(
            {
                "ok": True,
                "added": added,
                "cart_count": cart.get_item_count(),
                "message": (
                    f"דף ההזמנה עודכן — {added} פריטים נוספו"
                    if added
                    else "לא נוספו פריטים להזמנה"
                ),
            }
        )

    if added:
        messages.success(request, f"דף ההזמנה עודכן — {added} פריטים נוספו")
    else:
        messages.info(request, "לא נוספו פריטים להזמנה")

    return redirect(request.POST.get("next") or "white_catalog:cart")


@require_catalog_login
def cart_view(request):
    """Display the current cart contents."""
    context = {**_nav_context(request), **_build_cart_page_context(request)}
    return render(request, "white_catalog/cart.html", context)


@require_catalog_login
def cart_drawer(request):
    """HTML fragment for the cart side panel (AJAX)."""
    context = {
        **_nav_context(request),
        **_build_cart_page_context(request),
        "drawer_mode": True,
    }
    return render(request, "white_catalog/cart_drawer_ajax.html", context)


@require_POST
@require_catalog_login
def cart_update(request):
    """Update quantity or remove a single cart item."""
    user = get_current_catalog_user(request)
    item_id = request.POST.get("item_id")
    action = request.POST.get("action")  # "update" or "remove"
    variant_id = request.POST.get("variant_id")
    color_variant_id = request.POST.get("color_variant_id")
    pack_type_id = request.POST.get("pack_type_id")
    product_id = request.POST.get("product_id")
    is_ajax = request.POST.get("format") == "json"

    if settings.DEBUG:
        logger.warning(
            "cart_update start user=%s item_id=%s action=%s quantity=%s is_ajax=%s path=%s",
            getattr(user, "username", None),
            item_id,
            action,
            request.POST.get("quantity"),
            is_ajax,
            request.path,
        )

    item = None
    pack_type = None
    variant = None
    color_variant = None
    product = None
    if item_id:
        try:
            item = WhiteCartItem.objects.select_related("cart").get(pk=item_id, cart__user=user, cart__status=WhiteCart.STATUS_ACTIVE)
        except WhiteCartItem.DoesNotExist:
            if settings.DEBUG:
                logger.warning(
                    "cart_update missing item user=%s item_id=%s action=%s",
                    getattr(user, "username", None),
                    item_id,
                    action,
                )
            messages.error(request, "הפריט לא נמצא")
            return redirect("white_catalog:cart")
    elif action == "update" and variant_id and pack_type_id:
        cart = get_or_create_active_cart(user)
        try:
            variant = WhiteProductVariant.objects.select_related("product").get(
                pk=variant_id,
                is_active=True,
            )
            pack_type = WhitePackType.objects.get(pk=pack_type_id, is_active=True)
        except (WhiteProductVariant.DoesNotExist, WhitePackType.DoesNotExist, TypeError, ValueError):
            if is_ajax:
                return JsonResponse({"status": "error", "error": "הפריט לא נמצא"}, status=400)
            messages.error(request, "הפריט לא נמצא")
            return redirect("white_catalog:cart")

        # Enforce the user's pack route and the variant's pack forms.
        if pack_type.quantity != 3:
            if is_ajax:
                return JsonResponse({"status": "error", "error": "ביגוד מוזמן בשלישיות בלבד"}, status=400)
            messages.error(request, "ביגוד מוזמן בשלישיות בלבד")
            return redirect("white_catalog:cart")
        allowed_pack_ids = set(user.get_allowed_pack_types().values_list("id", flat=True))
        variant_pack_ids = set(variant.pack_types.values_list("id", flat=True))
        if pack_type.pk not in allowed_pack_ids or (variant_pack_ids and pack_type.pk not in variant_pack_ids):
            if is_ajax:
                return JsonResponse({"status": "error", "error": "סוג האריזה אינו זמין עבורך"}, status=400)
            messages.error(request, "סוג האריזה אינו זמין עבורך")
            return redirect("white_catalog:cart")
        item = (
            WhiteCartItem.objects.select_related("cart")
            .filter(cart=cart, variant=variant, pack_type=pack_type)
            .first()
        )
    elif action == "update" and color_variant_id:
        cart = get_or_create_active_cart(user)
        try:
            color_variant = WhiteColorVariant.objects.select_related("product").get(
                pk=color_variant_id,
                is_active=True,
            )
        except (WhiteColorVariant.DoesNotExist, TypeError, ValueError):
            if is_ajax:
                return JsonResponse({"status": "error", "error": "הפריט לא נמצא"}, status=400)
            messages.error(request, "הפריט לא נמצא")
            return redirect("white_catalog:cart")
        item = (
            WhiteCartItem.objects.select_related("cart")
            .filter(cart=cart, color_variant=color_variant)
            .first()
        )
    elif action == "update" and product_id:
        cart = get_or_create_active_cart(user)
        try:
            product = WhiteSubcategory.objects.get(pk=product_id, is_orderable=True, has_order_variants=False, has_color_variants=False)
        except (WhiteSubcategory.DoesNotExist, TypeError, ValueError):
            if is_ajax:
                return JsonResponse({"status": "error", "error": "הפריט לא נמצא"}, status=400)
            messages.error(request, "הפריט לא נמצא")
            return redirect("white_catalog:cart")
        item = (
            WhiteCartItem.objects.select_related("cart")
            .filter(cart=cart, product=product, variant__isnull=True, pack_type__isnull=True, color_variant__isnull=True)
            .first()
        )
    else:
        if is_ajax:
            return JsonResponse({"status": "error", "error": "הפריט לא נמצא"}, status=400)
        messages.error(request, "הפריט לא נמצא")
        return redirect("white_catalog:cart")

    if action == "remove":
        cart_obj = item.cart
        cart_pk = cart_obj.pk
        item.delete()
        if is_ajax:
            try:
                cart = WhiteCart.objects.get(pk=cart_pk)
                ct = "{:.2f}".format(cart.get_total())
                cc = cart.get_item_count()
            except WhiteCart.DoesNotExist:
                ct = "0.00"
                cc = 0
            if settings.DEBUG:
                logger.warning(
                    "cart_update removed item_id=%s cart_total=%s cart_count=%s",
                    item_id,
                    ct,
                    cc,
                )
            return JsonResponse({"status": "removed", "cart_total": ct, "cart_count": cc})
        messages.success(request, "הפריט הוסר מההזמנה")
    elif action == "update":
        try:
            qty = int(request.POST.get("quantity", 0))
        except ValueError:
            qty = 0
        if qty <= 0:
            if not item:
                cart_obj = get_or_create_active_cart(user)
                cart_pk = cart_obj.pk
                if is_ajax:
                    return JsonResponse({
                        "status": "noop",
                        "cart_total": "{:.2f}".format(cart_obj.get_total()),
                        "cart_count": cart_obj.get_item_count(),
                    })
                return redirect("white_catalog:cart")
            cart_obj = item.cart
            cart_pk = cart_obj.pk
            item.delete()
            if is_ajax:
                try:
                    cart = WhiteCart.objects.get(pk=cart_pk)
                    ct = "{:.2f}".format(cart.get_total())
                    cc = cart.get_item_count()
                except WhiteCart.DoesNotExist:
                    ct = "0.00"
                    cc = 0
                if settings.DEBUG:
                    logger.warning(
                        "cart_update qty<=0 removed item_id=%s cart_total=%s cart_count=%s",
                        item_id,
                        ct,
                        cc,
                    )
                return JsonResponse({"status": "removed", "cart_total": ct, "cart_count": cc})
            messages.success(request, "הפריט הוסר מההזמנה")
        else:
            if not item:
                if product:
                    effective_price = apply_price_list(product.unit_price, user)
                elif color_variant:
                    effective_price = apply_price_list(color_variant.get_effective_price(), user)
                else:
                    effective_price = apply_price_list(
                        variant.unit_price if variant.unit_price is not None else variant.product.unit_price,
                        user,
                    )
                if effective_price is None:
                    if is_ajax:
                        return JsonResponse({"status": "error", "error": "מחיר לא זמין"}, status=400)
                    messages.error(request, "מחיר לא זמין")
                    return redirect("white_catalog:cart")
                if product:
                    item = WhiteCartItem.objects.create(
                        cart=cart,
                        product=product,
                        quantity=qty,
                        price_at_add=effective_price,
                    )
                elif color_variant:
                    item, _ = WhiteCartItem.objects.update_or_create(
                        cart=cart,
                        color_variant=color_variant,
                        defaults={
                            "product": color_variant.product,
                            "quantity": qty,
                            "price_at_add": effective_price,
                        },
                    )
                else:
                    item, _ = WhiteCartItem.objects.update_or_create(
                        cart=cart,
                        variant=variant,
                        pack_type=pack_type,
                        defaults={
                            "product": variant.product,
                            "quantity": qty,
                            "price_at_add": effective_price * pack_type.quantity,
                        },
                    )
            else:
                item.quantity = qty
                item.save(update_fields=["quantity", "updated"])
            if is_ajax:
                line_total = item.price_at_add * qty
                cart_total = "{:.2f}".format(item.cart.get_total())
                cart_count = item.cart.get_item_count()
                if settings.DEBUG:
                    logger.warning(
                        "cart_update saved item_id=%s qty=%s line_total=%s cart_total=%s cart_count=%s",
                        item.id,
                        qty,
                        "{:.2f}".format(line_total),
                        cart_total,
                        cart_count,
                    )
                return JsonResponse({
                    "status": "ok",
                    "item_id": item.id,
                    "line_total": "{:.2f}".format(line_total),
                    "cart_total": cart_total,
                    "cart_count": cart_count,
                })

    return redirect("white_catalog:cart")


@require_POST
@require_catalog_login
def cart_clear(request):
    """Remove all items from the active cart."""
    user = get_current_catalog_user(request)
    WhiteCart.objects.filter(user=user, status=WhiteCart.STATUS_ACTIVE).delete()
    messages.success(request, "ההזמנה נוקתה בהצלחה")
    return redirect("white_catalog:cart")


# ---------------------------------------------------------------------------
# Checkout & Orders
# ---------------------------------------------------------------------------

@require_catalog_login
def checkout(request):
    """Submit the order (POST only). GET redirects to cart."""
    if request.method != "POST":
        return redirect("white_catalog:cart")

    user = get_current_catalog_user(request)
    try:
        cart = WhiteCart.objects.prefetch_related(
            "items__product", "items__variant__fabric_type", "items__variant__size_type",
            "items__color_variant__color", "items__pack_type"
        ).get(user=user, status=WhiteCart.STATUS_ACTIVE)
    except WhiteCart.DoesNotExist:
        messages.error(request, "דף ההזמנה ריק")
        return redirect("white_catalog:cart")

    if not cart.items.exists():
        messages.error(request, "דף ההזמנה ריק")
        return redirect("white_catalog:cart")

    notes = request.POST.get("notes", "").strip()
    total = cart.get_total()

    order = WhiteOrder.objects.create(
        user=user,
        cart=cart,
        status=WhiteOrder.STATUS_PENDING,
        notes=notes,
        total_amount=total,
    )

    for item in cart.items.select_related(
        "product", "variant__fabric_type", "variant__size_type", "color_variant__color", "pack_type"
    ).all():
        if item.color_variant_id:
            snapshot_barcode = item.color_variant.barcode or ""
        elif item.variant_id:
            snapshot_barcode = item.variant.barcode or ""
        else:
            snapshot_barcode = ""
        WhiteOrderItem.objects.create(
            order=order,
            product=item.product,
            variant=item.variant,
            color_variant=item.color_variant,
            product_name=item.product.name,
            variant_name="מניפת צבעים" if item.color_variant_id else item.display_variant_name,
            color_name=item.color_variant.color.name if item.color_variant_id else "",
            barcode=snapshot_barcode,
            size_name=item.display_size_name,
            pack_type_name=item.display_pack_name,
            pack_quantity=item.pack_type.quantity if item.pack_type_id else 1,
            quantity=item.quantity,
            unit_price=item.price_at_add,
        )

    cart.status = WhiteCart.STATUS_SUBMITTED
    cart.save(update_fields=["status", "updated"])

    messages.success(request, f"ההזמנה {order.order_number} התקבלה בהצלחה!")
    return redirect("white_catalog:order_confirm", order_number=order.order_number)


@require_catalog_login
def order_confirm(request, order_number):
    """Order confirmation page."""
    user = get_current_catalog_user(request)
    order = get_object_or_404(
        WhiteOrder.objects.prefetch_related(
            "items__product__images",
            "items__variant__fabric_type",
            "items__variant__size_type",
        ),
        order_number=order_number,
        user=user,
    )
    context = {
        **_nav_context(request),
        "order": order,
        "grouped_items": _build_order_grouped_items(order),
        "cart_count": _cart_count(request),
    }
    return render(request, "white_catalog/order_confirm.html", context)


@require_catalog_login
def order_list(request):
    """List of all orders for the logged-in user."""
    user = get_current_catalog_user(request)
    orders = WhiteOrder.objects.filter(user=user).prefetch_related("items").order_by("-created")
    context = {
        **_nav_context(request),
        "orders": orders,
        "cart_count": _cart_count(request),
    }
    return render(request, "white_catalog/order_list.html", context)


# ---------------------------------------------------------------------------
# Product data export (for customer site integration)
# ---------------------------------------------------------------------------

@require_catalog_login
def export_products_excel(request):
    """Download the full product catalog as an Excel file.

    One row per product variant (barcode + fabric + size) with wholesale and
    recommended retail prices, pack prices filtered by the user's pack route,
    marketing description and absolute image URLs — everything a customer
    needs to import the products into their own store.
    """
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter

    user = get_current_catalog_user(request)
    allowed_packs = list(user.get_allowed_pack_types().order_by("order", "quantity"))

    headers = [
        "ברקוד",
        "קטגוריה",
        "שם מוצר",
        "סוג בד",
        "מידה",
        "צבע",
        'מחיר סיטונאי ליחידה (לא כולל מע"מ)',
    ]
    headers += [f'מחיר {pack.name} (לא כולל מע"מ)' for pack in allowed_packs]
    headers += [
        'מחיר קמעונאי מומלץ (לא כולל מע"מ)',
        "תיאור שיווקי",
        "קישורי תמונות",
    ]
    price_columns = set(range(7, 7 + len(allowed_packs) + 2))  # unit + packs + retail
    images_column = len(headers)

    wb = Workbook()
    ws = wb.active
    ws.title = "קטלוג מוצרים"
    ws.sheet_view.rightToLeft = True

    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="7594B1", end_color="7594B1", fill_type="solid")
    for col, title in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col, value=title)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws.freeze_panes = "A2"

    widths = [16, 20, 28, 20, 12, 14] + [18] * (len(allowed_packs) + 2) + [60, 70]
    for col, width in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(col)].width = width

    def clean_html(value):
        return " ".join(strip_tags(value or "").split())

    def write_row(row_index, values):
        for col, value in enumerate(values, start=1):
            cell = ws.cell(row=row_index, column=col, value=value)
            if col in price_columns and value is not None:
                cell.number_format = "#,##0.00"
            if col == images_column:
                cell.alignment = Alignment(wrap_text=True, vertical="top")

    products = (
        WhiteSubcategory.objects.filter(is_orderable=True)
        .select_related("category")
        .prefetch_related(
            "images",
            "variants__fabric_type",
            "variants__size_type",
            "variants__pack_types",
            "variants__pack_prices",
            "color_variants__color",
        )
        .order_by("category__order", "category__name", "order", "name")
    )

    row = 2
    for product in products:
        category_name = product.category.name if product.category_id else ""
        description = clean_html(product.marketing_description) or clean_html(product.description)
        image_urls = "\n".join(
            request.build_absolute_uri(img["url"]) for img in product.get_all_images()
        )

        if product.has_order_variants:
            for variant in product.variants.all():
                if not variant.is_active:
                    continue
                # Empty pack_types on a variant means it is sold in all pack forms.
                variant_pack_ids = {pt.pk for pt in variant.pack_types.all()}
                pack_price_map = {pp.pack_type_id: pp.price for pp in variant.pack_prices.all()}
                effective_unit = apply_price_list(
                    variant.unit_price if variant.unit_price is not None else product.unit_price,
                    user,
                )

                pack_cells = []
                available_in_any_pack = False
                for pack in allowed_packs:
                    if variant_pack_ids and pack.pk not in variant_pack_ids:
                        pack_cells.append(None)
                        continue
                    raw_pack = pack_price_map.get(pack.pk)
                    if raw_pack is not None:
                        price = apply_price_list(raw_pack, user)
                    elif effective_unit is not None:
                        price = effective_unit * pack.quantity
                    else:
                        price = None
                    pack_cells.append(price)
                    available_in_any_pack = True

                if allowed_packs and not available_in_any_pack:
                    continue

                write_row(row, [
                    variant.barcode or "",
                    category_name,
                    product.name,
                    variant.fabric_type.name,
                    variant.size_type.name,
                    "",
                    effective_unit,
                    *pack_cells,
                    product.online_price,
                    description,
                    image_urls,
                ])
                row += 1
        elif product.has_color_variants:
            for color_variant in product.color_variants.all():
                if not color_variant.is_active:
                    continue
                color_image_urls = image_urls
                if color_variant.image:
                    color_image_urls = "\n".join(filter(None, [
                        request.build_absolute_uri(color_variant.image.url),
                        image_urls,
                    ]))
                write_row(row, [
                    color_variant.barcode or "",
                    category_name,
                    product.name,
                    "",
                    "",
                    color_variant.color.name,
                    apply_price_list(color_variant.get_effective_price(), user),
                    *[None] * len(allowed_packs),
                    product.online_price,
                    description,
                    color_image_urls,
                ])
                row += 1
        else:
            write_row(row, [
                "",
                category_name,
                product.name,
                "",
                "",
                "",
                apply_price_list(product.unit_price, user),
                *[None] * len(allowed_packs),
                product.online_price,
                description,
                image_urls,
            ])
            row += 1

    buffer = BytesIO()
    wb.save(buffer)

    filename = f"arye-white-catalog-{timezone.now().strftime('%Y-%m-%d')}.xlsx"
    response = HttpResponse(
        buffer.getvalue(),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    response["Content-Disposition"] = f'attachment; filename="{filename}"'
    return response


@require_catalog_login
def export_order_excel(request, order_number):
    """Download a single order as an Excel file.

    One row per order line with the variant barcode, quantities and prices —
    so the customer can receive the ordered stock into their own system
    without the noise of the full catalog.
    """
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter

    user = get_current_catalog_user(request)
    order = get_object_or_404(
        WhiteOrder.objects.prefetch_related("items__product__images", "items__variant"),
        order_number=order_number,
        user=user,
    )

    headers = [
        "ברקוד",
        "שם מוצר",
        "גרסה (בד)",
        "מידה",
        "סוג מארז",
        "יחידות במארז",
        "כמות מארזים",
        'סה"כ יחידות',
        'מחיר למארז (לא כולל מע"מ)',
        'סה"כ שורה (לא כולל מע"מ)',
        'מחיר קמעונאי מומלץ (לא כולל מע"מ)',
        "קישורי תמונות",
    ]
    price_columns = {9, 10, 11}
    images_column = len(headers)

    wb = Workbook()
    ws = wb.active
    ws.title = f"הזמנה {order.order_number}"[:31]
    ws.sheet_view.rightToLeft = True

    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="7594B1", end_color="7594B1", fill_type="solid")
    for col, title in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col, value=title)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws.freeze_panes = "A2"

    widths = [16, 28, 20, 12, 16, 12, 12, 12, 18, 18, 18, 70]
    for col, width in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(col)].width = width

    def write_row(row_index, values):
        for col, value in enumerate(values, start=1):
            cell = ws.cell(row=row_index, column=col, value=value)
            if col in price_columns and value is not None:
                cell.number_format = "#,##0.00"
            if col == images_column:
                cell.alignment = Alignment(wrap_text=True, vertical="top")

    row = 2
    for item in order.items.all():
        barcode = item.barcode or (item.variant.barcode if item.variant_id and item.variant else None)
        product = item.product if item.product_id else None
        online_price = product.online_price if product else None
        image_urls = "\n".join(
            request.build_absolute_uri(img["url"]) for img in product.get_all_images()
        ) if product else ""

        write_row(row, [
            barcode or "",
            item.product_name,
            item.variant_name,
            item.size_name,
            item.pack_type_name,
            item.pack_quantity,
            item.quantity,
            item.pack_quantity * item.quantity,
            item.unit_price,
            item.get_line_total(),
            online_price,
            image_urls,
        ])
        row += 1

    total_font = Font(bold=True)
    label_cell = ws.cell(row=row, column=9, value='סה"כ להזמנה (לא כולל מע"מ):')
    label_cell.font = total_font
    total_cell = ws.cell(row=row, column=10, value=order.total_amount)
    total_cell.font = total_font
    total_cell.number_format = "#,##0.00"

    buffer = BytesIO()
    wb.save(buffer)

    filename = f"arye-order-{order.order_number}.xlsx"
    response = HttpResponse(
        buffer.getvalue(),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    response["Content-Disposition"] = f'attachment; filename="{filename}"'
    return response

